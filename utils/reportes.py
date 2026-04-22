"""
utils/reportes.py
Generación de reportes en Excel y PDF en memoria (BytesIO).
Sin archivos temporales — compatible con Streamlit Cloud.
"""
from __future__ import annotations
import io
from datetime import datetime


# ─────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────

def _fmt(iso_str: str) -> str:
    if not iso_str:
        return "—"
    try:
        return datetime.fromisoformat(iso_str.replace("Z", "+00:00")).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return iso_str


def _hoy() -> str:
    return datetime.now().strftime("%d/%m/%Y %H:%M")


def _hoy_file() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M")


# ─────────────────────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────────────────────

def _excel_base(titulo: str, columnas: list[str], filas: list[list]) -> bytes:
    """Genera un Excel en memoria y lo retorna como bytes."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Instala openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Colores institucionales
    NAVY   = "0D2137"
    WHITE  = "FFFFFF"
    LIGHT  = "D0E8F7"
    STRIPE = "F0F5FA"

    # Título
    ws.merge_cells(f"A1:{get_column_letter(len(columnas))}1")
    cell_titulo = ws["A1"]
    cell_titulo.value         = titulo
    cell_titulo.font          = Font(name="Calibri", bold=True, size=14, color=WHITE)
    cell_titulo.fill          = PatternFill("solid", fgColor=NAVY)
    cell_titulo.alignment     = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtítulo con fecha
    ws.merge_cells(f"A2:{get_column_letter(len(columnas))}2")
    cell_fecha = ws["A2"]
    cell_fecha.value      = f"Generado: {_hoy()}  |  Instituto Tecnológico de Matehuala"
    cell_fecha.font       = Font(name="Calibri", italic=True, size=10, color=NAVY)
    cell_fecha.fill       = PatternFill("solid", fgColor=LIGHT)
    cell_fecha.alignment  = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Encabezados
    header_row = 3
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(columnas, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font      = Font(name="Calibri", bold=True, size=11, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor="163352")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[header_row].height = 20

    # Datos
    for row_idx, fila in enumerate(filas, start=header_row + 1):
        bg = STRIPE if row_idx % 2 == 0 else WHITE
        for col_idx, valor in enumerate(fila, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(valor) if valor is not None else "—")
            cell.font      = Font(name="Calibri", size=10)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[row_idx].height = 16

    # Ancho automático de columnas
    for col_idx, col_name in enumerate(columnas, start=1):
        letter = get_column_letter(col_idx)
        max_len = max(
            len(col_name),
            *[len(str(fila[col_idx-1])) for fila in filas] if filas else [0]
        )
        ws.column_dimensions[letter].width = min(max_len + 4, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────
# PDF
# ─────────────────────────────────────────────────────────

def _pdf_base(titulo: str, columnas: list[str], filas: list[list],
              subtitulo: str = "") -> bytes:
    """Genera un PDF en memoria y lo retorna como bytes."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        raise ImportError("Instala reportlab: pip install reportlab")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    NAVY   = colors.HexColor("#0D2137")
    BLUE   = colors.HexColor("#1A6FA8")
    LIGHT  = colors.HexColor("#D0E8F7")
    STRIPE = colors.HexColor("#F0F5FA")
    WHITE  = colors.white

    style_titulo = ParagraphStyle("titulo", fontSize=16, textColor=WHITE,
                                   alignment=TA_CENTER, fontName="Helvetica-Bold",
                                   spaceAfter=2)
    style_sub    = ParagraphStyle("sub", fontSize=9, textColor=NAVY,
                                   alignment=TA_CENTER, fontName="Helvetica",
                                   spaceAfter=6)

    story = []

    # Encabezado con fondo navy
    header_data = [[Paragraph(titulo, style_titulo)]]
    header_tbl  = Table(header_data, colWidths=[doc.width])
    header_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), NAVY),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 0.2*cm))

    sub_txt = subtitulo or f"Generado: {_hoy()}  |  Instituto Tecnológico de Matehuala"
    story.append(Paragraph(sub_txt, style_sub))
    story.append(Spacer(1, 0.3*cm))

    # Tabla de datos
    col_width = doc.width / len(columnas)
    tabla_data = [columnas] + [[str(v) if v is not None else "—" for v in fila]
                                for fila in filas]

    tbl = Table(tabla_data, colWidths=[col_width]*len(columnas), repeatRows=1)

    row_styles = [
        ("BACKGROUND",    (0,0), (-1,0), BLUE),
        ("TEXTCOLOR",     (0,0), (-1,0), WHITE),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0), 9),
        ("ALIGN",         (0,0), (-1,0), "CENTER"),
        ("TOPPADDING",    (0,0), (-1,0), 6),
        ("BOTTOMPADDING", (0,0), (-1,0), 6),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 8),
        ("ALIGN",         (0,1), (-1,-1), "LEFT"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,1), (-1,-1), 4),
        ("BOTTOMPADDING", (0,1), (-1,-1), 4),
        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [WHITE, STRIPE]),
    ]
    tbl.setStyle(TableStyle(row_styles))
    story.append(tbl)

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────
# REPORTES PÚBLICOS (llamados desde las páginas)
# ─────────────────────────────────────────────────────────

def reporte_alumno_excel(sesiones: list[dict], nombre_alumno: str) -> tuple[bytes, str]:
    cols  = ["Fecha y Hora", "Docente", "Materia", "Estado", "Asistencia", "Notas"]
    filas = []
    for s in sesiones:
        asist = ("Sí" if s.get("asistencia") is True
                 else "No" if s.get("asistencia") is False else "—")
        filas.append([
            _fmt(s.get("fecha_hora","")),
            s.get("docente_nombre","—"),
            s.get("materia","—"),
            s.get("estado","—"),
            asist,
            s.get("notas_docente","—") or "—",
        ])
    titulo   = f"Historial de Tutorías — {nombre_alumno}"
    filename = f"historial_{nombre_alumno.replace(' ','_')}_{_hoy_file()}.xlsx"
    return _excel_base(titulo, cols, filas), filename


def reporte_alumno_pdf(sesiones: list[dict], nombre_alumno: str) -> tuple[bytes, str]:
    cols  = ["Fecha y Hora", "Docente", "Materia", "Estado", "Asistencia"]
    filas = []
    for s in sesiones:
        asist = ("Sí" if s.get("asistencia") is True
                 else "No" if s.get("asistencia") is False else "—")
        filas.append([
            _fmt(s.get("fecha_hora","")),
            s.get("docente_nombre","—"),
            s.get("materia","—"),
            s.get("estado","—"),
            asist,
        ])
    titulo   = f"Historial de Tutorías — {nombre_alumno}"
    filename = f"historial_{nombre_alumno.replace(' ','_')}_{_hoy_file()}.pdf"
    return _pdf_base(titulo, cols, filas), filename


def reporte_docente_excel(sesiones: list[dict], nombre_docente: str) -> tuple[bytes, str]:
    cols  = ["Fecha y Hora", "Alumno", "No. Control", "Materia", "Estado", "Asistencia", "Notas"]
    filas = []
    for s in sesiones:
        asist = ("Sí" if s.get("asistencia") is True
                 else "No" if s.get("asistencia") is False else "—")
        filas.append([
            _fmt(s.get("fecha_hora","")),
            s.get("alumno_nombre","—"),
            s.get("alumno_control","—"),
            s.get("materia","—"),
            s.get("estado","—"),
            asist,
            s.get("notas_docente","—") or "—",
        ])
    titulo   = f"Historial de Tutorías — {nombre_docente}"
    filename = f"historial_{nombre_docente.replace(' ','_')}_{_hoy_file()}.xlsx"
    return _excel_base(titulo, cols, filas), filename


def reporte_docente_pdf(sesiones: list[dict], nombre_docente: str) -> tuple[bytes, str]:
    cols  = ["Fecha y Hora", "Alumno", "No. Control", "Materia", "Estado", "Asistencia"]
    filas = []
    for s in sesiones:
        asist = ("Sí" if s.get("asistencia") is True
                 else "No" if s.get("asistencia") is False else "—")
        filas.append([
            _fmt(s.get("fecha_hora","")),
            s.get("alumno_nombre","—"),
            s.get("alumno_control","—"),
            s.get("materia","—"),
            s.get("estado","—"),
            asist,
        ])
    titulo   = f"Historial de Tutorías — {nombre_docente}"
    filename = f"historial_{nombre_docente.replace(' ','_')}_{_hoy_file()}.pdf"
    return _pdf_base(titulo, cols, filas), filename


def reporte_admin_excel(sesiones: list[dict],
                        filtro: str = "General") -> tuple[bytes, str]:
    cols  = ["Fecha y Hora", "Alumno", "No. Control",
             "Docente", "Depto.", "Materia", "Estado", "Asistencia"]
    filas = []
    for s in sesiones:
        asist = ("Sí" if s.get("asistencia") is True
                 else "No" if s.get("asistencia") is False else "—")
        filas.append([
            _fmt(s.get("fecha_hora","")),
            s.get("alumno_nombre","—"),
            s.get("alumno_control","—"),
            s.get("docente_nombre","—"),
            s.get("docente_departamento","—"),
            s.get("materia","—"),
            s.get("estado","—"),
            asist,
        ])
    titulo   = f"Reporte Global de Tutorías — {filtro}"
    filename = f"reporte_global_{filtro.replace(' ','_')}_{_hoy_file()}.xlsx"
    return _excel_base(titulo, cols, filas), filename


def reporte_admin_pdf(sesiones: list[dict],
                      filtro: str = "General") -> tuple[bytes, str]:
    cols  = ["Fecha y Hora", "Alumno", "No. Control",
             "Docente", "Materia", "Estado", "Asistencia"]
    filas = []
    for s in sesiones:
        asist = ("Sí" if s.get("asistencia") is True
                 else "No" if s.get("asistencia") is False else "—")
        filas.append([
            _fmt(s.get("fecha_hora","")),
            s.get("alumno_nombre","—"),
            s.get("alumno_control","—"),
            s.get("docente_nombre","—"),
            s.get("materia","—"),
            s.get("estado","—"),
            asist,
        ])
    titulo   = f"Reporte Global de Tutorías — {filtro}"
    filename = f"reporte_global_{filtro.replace(' ','_')}_{_hoy_file()}.pdf"
    return _pdf_base(titulo, cols, filas), filename
